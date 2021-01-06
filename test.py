
'''

from openpyxl import Workbook, load_workbook
import datetime


book = Workbook()
sheet = book.active

sheet.cell(1, 1).value = 1

book.save("diet_diary2.xlsx")
'''


a = input("Test")
while not a.isdigit():
    print("Enter a number")
    break
