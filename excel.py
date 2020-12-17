import json

from openpyxl import Workbook, load_workbook
import datetime
from globals import get_global
from json_writer import json_load


def excel_exist():
    from os import path
    if not path.exists("diet_diary.xlsx"):
        book = Workbook()
        sheet = book.active
        sheet['A1'] = "Date"
        sheet['B1'] = "Total Calories consumed"
        sheet['C1'] = "Net+-"
        sheet['D1'] = "Weight"
        sheet['E1'] = "Daily Calories: 1600"
        book.save("diet_diary.xlsx")
        print("diet_diary.xlsx doesnt exit. It has been created")
    else:
        print("Excel File exists.")


def excel_write_test(calories_add):
    book = load_workbook("diet_diary.xlsx")
    sheet = book.active
    a = 2
    target_row = json_load()
    # target_row = get_global("target_row")
    if target_row == 0:
        target_row = check_today_row(sheet, a)

    if target_row is not None:
        if sheet.cell(target_row, 2).value is None:
            sheet.cell(target_row, 2).value = 0
        sheet.cell(target_row, 2).value += calories_add
        sheet.cell(target_row, 3).value = sheet.cell(target_row, 2).value - get_global("calorie_goal")

    book.save("diet_diary.xlsx")


def read_cell(cell_row, cell_col):
    book = load_workbook("diet_diary.xlsx")
    sheet = book.active
    return sheet.cell(cell_row, cell_col).value


def delete_cell(cell):
    book = load_workbook("diet_diary.xlsx")
    sheet = book.active
    sheet[cell].value = None
    book.save("diet_diary.xlsx")


def write_today_date_to_excel():
    book = load_workbook("diet_diary.xlsx")
    sheet = book.active
    a = 2

    while True:
        if sheet.cell(a, 1).value is None:
            sheet.cell(a, 1).value = datetime.date.today()
            break
        else:
            a += 1
            if a > 100:
                break
    book.save("diet_diary.xlsx")


def check_today_row(sheet, a):
    while True:
        if sheet.cell(a, 1).value is not None and sheet.cell(a, 2).value is None:
            with open("target_row.json", "r") as read:
                load_data = json.load(read)
                with open("target_row.json", "w") as write:
                    load_data["target_row"] = a
                    json.dump(load_data, write, indent=4)
                    write.close()
            return a

        elif sheet.cell(a, 1).value is None and sheet.cell(a, 2).value is None:
            print("Today's date is not set. Please press Start of the day Button")
            break
        else:
            a += 1
            if a > 20:
                break


def write_weight(q_dialog_input):
    book = load_workbook("diet_diary.xlsx")
    sheet = book.active
    if sheet.cell(json_load(), 4).value is None:
        sheet.cell(json_load(), 4).value = q_dialog_input
    else:
        print("Wrong day setting or already noted the weight for today. If wrong day: press Start of the day first")
    book.save("diet_diary.xlsx")
